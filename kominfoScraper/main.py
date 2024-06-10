from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from InquirerPy.base.control import Choice
import undetected_chromedriver as uc
from InquirerPy import inquirer
from selenium import webdriver
import pandas as pd
import openpyxl,os,time,colorama
from colorama import init,Fore


init(autoreset=True)
class Hoax:
    def __init__(self):
        option = uc.ChromeOptions() 
        # option.add_argument('--disable-gpu')
        # option.add_argument('--headless')
        self.driver = uc.Chrome(options=option, use_subprocess=True)
        self.driver.set_page_load_timeout(30)
        self.outWorkbook = openpyxl.load_workbook("HoaxData.xlsx")
        self.outSheet = self.outWorkbook.active
        self.outSheet["A1"] = "Title"
        self.outSheet["B1"] = "Category"
        self.outSheet["C1"] = "Url"
        self.outSheet["D1"] = "Image"
        self.outSheet["E1"] = "Author"
        self.outSheet["F1"] = "Dekripsi"
        self.outSheet["G1"] = "Counter Link"
        self.outSheet["H1"] = "Date"
        self.outSheet["I1"] = "Page"

    def GetCurrentData(self):
        try:
            df = pd.read_excel("HoaxData.xlsx", engine='openpyxl')
            lastRow = df.index[-1] + 2
            lastPage = int(df['Page'].dropna().tolist()[-1])
            return lastRow, lastPage
        except Exception as e:
            print(e)
            return 1, 1

    def Setup(self, page):
        try:
            self.driver.get(f"https://www.kominfo.go.id/content/all/laporan_isu_hoaks?page={page}")
            time.sleep(2)
            if "You are now in line" in self.driver.page_source:
                while True:
                    if "You are now in line" not in self.driver.page_source:
                        break
                    print("[+] waiting")
        except Exception as e:
            print(Fore.RED + str(e))

    def GetArticle(self):
        try:
            elements = WebDriverWait(self.driver, 40).until(EC.presence_of_all_elements_located((By.XPATH, "//a[@class='title']")))
            return [x.text for x in elements], [x.get_attribute('href') for x in elements]
        except Exception as e:
            print(Fore.RED + f"error while get Article! the message: {str(e)}\n","output will be None")
            Fore.RESET
            return ["None"],["None"]

    def GetDate(self):
        try:
            if WebDriverWait(self.driver, 20).until(EC.url_contains("https://www.kominfo.go.id/content/detail")):
                elements = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[8]/div/div[2]/div/div[1]/div[1]/div[1]/div[1]")))
                print("date: ",elements.text.split("\n"))
                return elements.text.split('\n')[1]
        except Exception as e:
            print(Fore.YELLOW + f"error while get Date! the message: {str(e)}\n","output will be None")
            return None

    def GetAuthor(self, url):
        try:
            self.driver.get(url)
            if WebDriverWait(self.driver, 20).until(EC.url_contains("https://www.kominfo.go.id/content/detail")):
                Adetail = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[8]/div/div[2]/div/div[1]/div[1]/div[2]/div[1]")))
                print(Adetail.text.split("Kategori ")[1].split(" | ")[0])
                print(Adetail.text.split(" | ")[-1])
                return Adetail.text.split("Kategori ")[1].split(" | ")[0],Adetail.text.split(" | ")[-1]
        except Exception as e:
            print(Fore.YELLOW + f"error while get Author! the message: {str(e)}\n","output will be None")
            return None,None
        
    def GetDesc(self):
        try:
            if WebDriverWait(self.driver, 20).until(EC.url_contains("https://www.kominfo.go.id/content/detail")):   
                deskirpsi = WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='youtube-container']//p")))
                desc = [desc.text for desc in deskirpsi][1:][:-1][:-1]
                return desc
        except Exception as e:
            print(Fore.YELLOW + f"error while get Description! the message: {str(e)}\n","output will be None")
            return None
        
    def GetLinkC(self):
        self.driver.get("https://www.kominfo.go.id/content/detail/54418/hoaks-akun-instagram-mengatasnamakan-pt-djarum/0/laporan_isu_hoaks")
        try:
            if WebDriverWait(self.driver, 20).until(EC.url_contains("https://www.kominfo.go.id/content/detail")):   
                Lcounter = WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='youtube-container']//ul//a")))
                lCounters = [lCounter.get_attribute('href') for lCounter in Lcounter]
                return lCounters
        except Exception as e:
            try:
                Lcounter = WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='O0']//u//a")))
                lCounters = [lCounter.get_attribute('href') for lCounter in Lcounter]
                return lCounters
            except Exception as e:
                print(Fore.YELLOW + f"error while get Counter Link! the message: {str(e)}\n","output will be None")
                return ['None']

    def GetImage(self):
        try:
            img = WebDriverWait(self.driver, 20).until(EC.presence_of_all_elements_located((By.XPATH, "//img[@class='thumbnail-img artikel--bg-size-cover']")))
            imgu = [x.get_attribute('src') for x in img]
            return imgu
        except Exception as e:
            print(Fore.YELLOW + f"error while get Image! the message: {str(e)}\n","output will be None")
            return None

    def Main(self,pages):
        lastRow, lastPage = self.GetCurrentData()
        for x in range(lastPage+1, lastPage + pages+1):
            self.Setup(x)
            title, url = self.GetArticle()
            print(Fore.BLUE + "[+] page: ",x)
            for index, page in enumerate(url):
                self.outSheet.cell(row=lastRow + 1, column=1, value=title[index])
                self.outSheet.cell(row=lastRow + 1, column=3, value=page)
                self.outSheet.cell(row=lastRow + 1, column=9, value=x)
                img = self.GetImage()   
                self.outSheet.cell(row=lastRow + 1, column=4, value=img[0])
                category,author = self.GetAuthor(page)
                desc = self.GetDesc()
                LinkC = self.GetLinkC()
                date = self.GetDate()
                self.outSheet.cell(row=lastRow + 1, column=8, value=date)
                self.outSheet.cell(row=lastRow + 1, column=6, value=str(desc))
                self.outSheet.cell(row=lastRow + 1, column=2, value=category)
                self.outSheet.cell(row=lastRow + 1, column=5, value=author)
                lastRow = lastRow + 1
                for index, linkC in enumerate(LinkC):
                    self.outSheet.cell(row=lastRow + 1, column=7, value=linkC)
                    lastRow = lastRow + 1
                lastRow = lastRow + 1
            print("\n")
        self.outWorkbook.save("HoaxData.xlsx")
        print(Fore.GREEN + "[+] done!")
        self.driver.quit()

class Satker:
    def __init__(self):
        option = uc.ChromeOptions() 
        # option.add_argument('--disable-gpu')
        # option.add_argument('--headless')
        self.driver = uc.Chrome(options=option, use_subprocess=True)
        self.driver.set_page_load_timeout(30)
        self.outWorkbook = openpyxl.load_workbook("SatkerData.xlsx")
        self.outSheet = self.outWorkbook.active
        self.outSheet["A1"] = "Title"
        self.outSheet["B1"] = "Category"
        self.outSheet["C1"] = "Url"
        self.outSheet["D1"] = "Image"
        self.outSheet["E1"] = "Author"
        self.outSheet["F1"] = "Deskripsi"
        self.outSheet["G1"] = "Date"
        self.outSheet["H1"] = "Page"

    def GetCurrentData(self):
        try:
            df = pd.read_excel("SatkerData.xlsx", engine='openpyxl')
            lastRow = df.index[-1] + 2
            lastPage = int(df['Page'].dropna().tolist()[-1])
            return lastRow, lastPage
        except Exception as e:
            print(e)
            return 1, 1

    def Setup(self, page):
        try:
            self.driver.get(f"https://www.kominfo.go.id/content/all/berita_satker?page={page}")
            # WebDriverWait(self.driver, 40).until(EC.((By.XPATH, "//a[@class='title']")))
            if "Request unsuccessful." in self.driver.page_source:
                while True:
                    if "You are now in line" not in self.driver.page_source:
                        break
                    print("waiting")
        except Exception as e:
            print(e)

    def GetArticle(self): #get title and news url
        try:
            elements = WebDriverWait(self.driver, 20).until(EC.presence_of_all_elements_located((By.XPATH, "//a[@class='title']")))
            return [x.text for x in elements], [x.get_attribute('href') for x in elements]
        except Exception as e:
            print(Fore.RED + f"error while get Article! the message: {str(e)}\n","output will be None")
            

    def GetDate(self): #get news date
        try:
            # self.driver.get("https://www.kominfo.go.id/content/detail/40367/dorong-inklusi-keuangan-daerah-kominfo-latih-warga-banjarbaru-kelola-keuangan-digital/0/berita_satker")
            if WebDriverWait(self.driver, 20).until(EC.url_contains("https://www.kominfo.go.id/content/detail")):
                elements = WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[8]/div/div[2]/div/div[1]/div[1]/div[1]/div[1]")))
                print("date: ",elements.text.split("\n"))
                return elements.text.split('\n')[1]
        except Exception as e:
            print(Fore.YELLOW + f"error while get Date! the message: {str(e)}\n","output will be None")
            return None

    def GetAuthor(self, url): #get news author
        try:
            self.driver.get(url)
            if WebDriverWait(self.driver, 20).until(EC.url_contains("https://www.kominfo.go.id/content/detail")):
                Adetail = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[8]/div/div[2]/div/div[1]/div[1]/div[2]/div[1]")))
                print(Adetail.text.split("Kategori ")[1].split(" | ")[0])
                print(Adetail.text.split(" | ")[-1])
                return Adetail.text.split("Kategori ")[1].split(" | ")[0],Adetail.text.split(" | ")[-1]
        except Exception as e:
            print(Fore.YELLOW + f"error while get Author! the message: {str(e)}\n","output will be None")
            return None,None

    def GetDesc(self):
        self.driver.get("https://www.kominfo.go.id/content/detail/55732/serpihan-logam-dalam-makanan-bayi-awas-hoaks/0/berita_satker")
        finalResult = []
        try:
            if WebDriverWait(self.driver, 30).until(EC.url_contains("https://www.kominfo.go.id/content/detail")):
                deskripsi = WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='youtube-container']//p")))
                for desc in deskripsi:
                    if desc.text != '':
                        finalResult.append(desc.text)
                print(finalResult[0])
                return finalResult 
        except Exception as e:
            print(Fore.YELLOW + f"error while get Description! the message: {str(e)}\n","output will be None")
            return ['None']

    def GetImage(self): #get news image
        try:
            img = WebDriverWait(self.driver, 20).until(EC.presence_of_all_elements_located((By.XPATH, "//img[@class='thumbnail-img artikel--bg-size-cover']")))
            imgu = [x.get_attribute('src') for x in img]
            return imgu
        except Exception as e:
            print(Fore.YELLOW + f"error while get Image! the message: {str(e)}\n","output will be None")
            return None
    def Main(self,pages):
        lastRow, lastPage = self.GetCurrentData()
        for x in range(lastPage+1, lastPage + pages+1):
            self.Setup(x)
            title, url = self.GetArticle()
            print(Fore.BLUE + "[+] page: ",x)
            for index, page in enumerate(url):
                self.outSheet.cell(row=lastRow + 1, column=1, value=title[index])
                self.outSheet.cell(row=lastRow + 1, column=3, value=page)
                self.outSheet.cell(row=lastRow + 1, column=8, value=x)
                img = self.GetImage()   
                self.outSheet.cell(row=lastRow + 1, column=4, value=img[0])
                category,author = self.GetAuthor(page)
                date = self.GetDate()
                desc = self.GetDesc()
                self.outSheet.cell(row=lastRow + 1, column=6, value=str(desc))
                self.outSheet.cell(row=lastRow + 1, column=7, value=date)
                self.outSheet.cell(row=lastRow + 1, column=2, value=category)
                self.outSheet.cell(row=lastRow + 1, column=5, value=author)
                lastRow+=1
            print("\n")
        self.outWorkbook.save("SatkerData.xlsx")
        print(Fore.GREEN + "[+] done!")
        self.driver.quit()

class UI:
    def __init__(self):
        self.proceed= False

    def main(self):
        action = inquirer.select(
            message="Select an action:",
            choices=[
                "Isu Hoax",
                "Satker",
                Choice(value=None, name="Exit"),
            ],
            default=None,
        ).execute()

        if action == "Isu Hoax":
            hox = Hoax()
            page = inquirer.text(
                message="insert jumlah page yang ingin di scrape: ",
                multicolumn_complete=True,
            ).execute()
            hox.Main(int(page))
        elif action == "Satker":
            satker = Satker()
            page = inquirer.text(
                message="insert jumlah page yang ingin di scrape: ",
                multicolumn_complete=True,
            ).execute()
            satker.Main(int(page))
if __name__ == "__main__":
    os.system('cls')
    myui = UI()
    myui.main()

# sat = Satker()
# sat.GetDesc()
